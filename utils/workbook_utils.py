import subprocess
import sys

def install_openpyxl():
    """
        Verifica se a biblioteca 'openpyxl' está instalada e a instala silenciosamente, se necessário.

        Esse tipo de verificação é útil quando o script precisa rodar em ambientes onde a dependência
        pode não estar previamente instalada.
    """
    try:
        import openpyxl
    except ImportError:
        # Executa um subprocesso para instalar o pacote usando o pip e o Python atual
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])

# Garante que openpyxl esteja disponível antes de importar
install_openpyxl()

from openpyxl import load_workbook

def load_workbooks(source_path, destination_path):
    """
        Carrega as planilhas Excel de origem e destino com tratamento de exceções.

        Args:
            source_path (str): Caminho completo para o arquivo de origem (.xlsx).
            destination_path (str): Caminho completo para o arquivo de destino (.xlsx).

        Returns:
            tuple: Uma tupla contendo dois objetos `Workbook` (origem, destino).

        Raises:
            FileNotFoundError: Se algum dos arquivos não for encontrado.
            PermissionError: Se houver problemas de permissão ao acessar os arquivos.
            ValueError: Para qualquer outro erro inesperado ao carregar os arquivos.
    """
    try:
        # Carrega a planilha de origem com valores calculados (sem fórmulas)
        source_workbook = load_workbook(source_path, data_only=True)

        # Carrega a planilha de destino com todas as fórmulas e dados
        destination_workbook = load_workbook(destination_path)

        return source_workbook, destination_workbook

    except FileNotFoundError as e:
        raise FileNotFoundError(f"Erro ao abrir os arquivos: {e}")
    except PermissionError as e:
        raise PermissionError(f"Permissão negada ao acessar os arquivos: {e}")
    except Exception as e:
        raise ValueError(f"Erro ao carregar as planilhas: {e}")
